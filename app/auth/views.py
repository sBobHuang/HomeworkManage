from flask import render_template, redirect, request, url_for, flash, current_app
from flask_login import current_user
from flask_login import login_user, logout_user, login_required
from flask_moment import datetime
from .forms import *
from ..model import *

from ..fuc import *
from . import auth
from ..main.export import exportOneHomeWorks
import os
from .. import db
from openpyxl import load_workbook
from sqlalchemy import func
from datetime import timedelta
from dateutil.relativedelta import relativedelta


@auth.before_app_request
def before_request():
    if current_user.is_authenticated:
        current_user.ping()


@auth.route('/login', methods=['GET', 'POST'])
# 登录界面
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is not None and user.verify_password(form.password.data):
            login_user(user, remember=True)
            return redirect(request.args.get('next') or url_for('main.index'))
        flash('用户名或者密码错误')
    return render_template('auth/login.html', form=form)


@auth.route('/logout')
# 退出
@login_required
def logout():
    logout_user()
    flash('您已经退出账号!')
    return redirect(url_for('main.index'))


@auth.route('/newUser', methods=['GET', 'POST'])
@login_required
def register():
    form = RegistrationForm()
    if not current_user.is_administrator():
        return redirect(url_for('main.index'))
    if form.validate_on_submit():
        if add_new_user(form.username.data, form.password2.data, '管理员注册'):
            flash('您已注册新用户!')
            return redirect(url_for('main.index'))
        else:
            flash("此账号已注册")
    return render_template('auth/register.html', form=form)


@auth.route('/admin', methods=['GET', 'POST'])
@login_required
def admin():
    current_term = current_app.config['CURRENT_TERM']
    if current_user.is_administrator() is False:
        return redirect(url_for('main.index'))
    query_term = request.args.get('query_term', current_term, type=str)
    add_course_form = AddCourseForm()
    add_courses_form = AddCoursesForm()
    if add_course_form.validate_on_submit():
        flash(add_course_form.course_name.data + '已添加')
        addCourseName(add_course_form.course_name.data)
    course_names = getCourseNames()
    if len(course_names) > 0:
        add_course_form.form_body = '已有科目：' + "、".join(course_names)
        add_courses_form.course.choices = [(i, i) for i in course_names]
    else:
        add_course_form.form_body = '暂无科目'
    if add_courses_form.validate_on_submit():
        flash(add_courses_form.course.data + '新班级已添加')
        addCourseNames(add_courses_form.course.data)
    forms = [add_course_form, add_courses_form]
    currentCourseInfos = CourseInfo.query.filter_by(course_period=query_term, disabled=False).order_by(
        CourseInfo.course_names).all()
    statusLabels = ['课程名称', '班级', '班级人数', '正在收的作业']
    statusContent = [CourseInfo.showStatus(i) for i in currentCourseInfos]
    return render_template('auth/admin.html',
                           statuslabels=statusLabels, statusContent=statusContent,
                           query_term=query_term,
                           query_term_str=courseInfoIDToStr(query_term),
                           forms=forms)


@auth.route('/courseManage', methods=['GET', 'POST'])
@login_required
def courseManage():
    page = request.args.get('page', 1, type=int)
    current_term = current_app.config['CURRENT_TERM']
    query_term = request.args.get('query_term', current_term, type=str)
    new_homework = request.args.get('new_homework', None, type=bool)
    pagination = CourseInfo.query.filter_by(course_period=query_term, disabled=False).order_by(
        CourseInfo.course_names).paginate(page, per_page=1)
    course_names = pagination.items[0].course_names
    form = UploadCoursesStus()
    if form.validate_on_submit():
        basedir = current_app.config['BASE_DIR']
        filename = form.file.data.filename
        file = os.path.splitext(filename)
        filename, file_type = file
        save_filename1 = '班级名单上传-' + course_names + file_type
        file_dir = os.path.join(basedir, "Upload")  # 拼接成合法文件夹地址
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)  # 文件夹不存在就创建
        save_filename = os.path.join(file_dir, save_filename1)
        form.file.data.save(save_filename)
        wb = load_workbook(save_filename)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for row in ws.rows:
            # print(row[0].value, row[1].value, row[2].value)
            student = Student(
                id=row[0].value,
                real_name=row[1].value,
                class_name=row[2].value,
            )
            student.course_names = course_names
            db.session.add(student)
            try:
                db.session.commit()
            except:
                db.session.rollback()
                print(row[1].value, '已插入')
    if new_homework is not None and new_homework is True:
        course = CourseInfo.query.filter_by(course_names=course_names).first()
        homework_dic = extendedInfoToDic(course.extended_info)
        # 并且不为空
        if course.extended_info is None or not bool(homework_dic):
            max_homework_id = 0
        else:
            max_homework_id = max([int(i) for i in homework_dic.keys()])
        course.extended_info = extendedInfoAdd(course.extended_info, max_homework_id + 1, 1)
    homework_del_id = request.args.get('homework_del_id', None, type=str)
    course = CourseInfo.query.filter_by(course_names=course_names).first()
    if homework_del_id is not None:
        course.extended_info = extendedInfoDel(course.extended_info, homework_del_id)
    homework_pause_id = request.args.get('homework_pause_id', None, type=str)
    if homework_pause_id is not None:
        course.extended_info = extendedInfoAdd(course.extended_info, homework_pause_id, 0)
    homework_continue_id = request.args.get('homework_continue_id', None, type=str)
    if homework_continue_id is not None:
        course.extended_info = extendedInfoAdd(course.extended_info, homework_continue_id, 1)
    student_query = Student.query.filter_by(course_names=course_names).all()
    export = request.args.get('export', None, type=bool)
    homework_export_id = request.args.get('homework_export_id', None, type=str)
    if export is not None and export is True > 0:
        if homework_export_id is not None:
            return exportOneHomeWorks(course.course_name, course_names, '作业' + homework_export_id)
    homeWorkManageLabels = ['作业名称', '删除', '暂停接收', '已交作业人数', '打包下载']
    courseManageLabels = ['序号', '学号', '姓名', '班级']
    homeWorkContent = homeWorkShow(course_names)
    for homework in homeWorkContent:
        courseManageLabels.append('作业' + homework + '上交时间')
    return render_template('auth/course_manage.html',
                           course=pagination.items[0].course_names,
                           homeWorkManageLabels=homeWorkManageLabels,
                           courseManageLabels=courseManageLabels,
                           courseContent=courseManageShow(student_query, homeWorkContent),
                           homeWorkContent=homeWorkContent,
                           pagination=pagination,
                           query_term=query_term,
                           form=form)


@auth.route('/file_upload', methods=['POST', 'GET'])
def file_upload():
    form = UploadForm()
    if form.validate_on_submit():
        basedir = current_app.config['BASE_DIR']
        file_dir = os.path.join(basedir, "File")  # 拼接成合法文件夹地址
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)  # 文件夹不存在就创建
        filename = form.file.data.filename
        files_query = FileRecord.query.filter_by(file_name=os.path.join(file_dir, filename)).first()
        if files_query is None:
            form.file.data.save(os.path.join(file_dir, filename))
            flash('作业上传成功')
            fileRecord = FileRecord(
                course_names='file_upload',
                real_name=filename,
                file_name=os.path.join(file_dir, filename)
            )
            db.session.add(fileRecord)
        else:
            print(files_query.file_name)
            os.remove(files_query.file_name)
            files_query.file_name = os.path.join(file_dir, filename)
            files_query.created_at = datetime.utcnow()
            flash("您已成功覆盖上次提交")
            form.file.data.save(os.path.join(file_dir, filename))
        db.session.commit()
        # if not os.path.exists(os.path.join(file_dir, new_filename)):
    return render_template('auth/file_du/upload.html', form=form)


@auth.route('/file_download', methods=['GET', 'POST'])
@login_required
def file_download():
    del_id = request.args.get('del_id', None, type=str)
    files_query = FileRecord.query.filter_by(id=del_id).first()
    if files_query is not None:
        db.session.delete(files_query)
        try:
            db.session.commit()
        except:
            db.session.rollback()
        os.remove(files_query.file_name)
    courses = FileRecord.query.filter_by(course_names='file_upload').order_by(FileRecord.created_at.desc()).all()
    courseManageLabels = ['编号', '文件名', '上传时间']
    courseContent = []
    for course in courses:
        courseContent.append(
            [course.id,
             course.real_name,
             course.created_at.strftime("%m/%d")
             ]
        )
    return render_template('auth/file_du/download.html',
                           courseManageLabels=courseManageLabels,
                           courseContent=courseContent
                           )


@auth.route('/oi_upload', methods=['POST', 'GET'])
def oi_upload():
    form = UploadForm()
    if form.validate_on_submit():
        basedir = current_app.config['BASE_DIR']
        file_dir = os.path.join(basedir, "OI")  # 拼接成合法文件夹地址
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)  # 文件夹不存在就创建
        filename = form.file.data.filename
        files_query = FileRecord.query.filter_by(file_name=os.path.join(file_dir, filename)).first()
        if files_query is None:
            form.file.data.save(os.path.join(file_dir, filename))
            flash('作业上传成功')
            fileRecord = FileRecord(
                course_names='oi_upload',
                real_name=filename,
                file_name=os.path.join(file_dir, filename)
            )
            db.session.add(fileRecord)
        else:
            print(files_query.file_name)
            os.remove(files_query.file_name)
            files_query.file_name = os.path.join(file_dir, filename)
            files_query.created_at = datetime.utcnow()
            flash("您已成功覆盖上次提交")
            form.file.data.save(os.path.join(file_dir, filename))
        db.session.commit()
        # if not os.path.exists(os.path.join(file_dir, new_filename)):
    return render_template('auth/file_du/upload.html', form=form)


@auth.route('/oi_download', methods=['GET', 'POST'])
def oi_download():
    courseManageLabels = ['编号', '文件名', '上传时间']
    courseContent = []
    del_id = request.args.get('del_id', None, type=str)
    files_query = FileRecord.query.filter_by(id=del_id).first()
    if files_query is not None:
        db.session.delete(files_query)
        try:
            db.session.commit()
        except:
            db.session.rollback()
        os.remove(files_query.file_name)
    courses = FileRecord.query.filter_by(course_names='oi_upload').order_by(FileRecord.created_at.desc()).all()
    for course in courses:
        courseContent.append(
            [course.id,
             course.real_name,
             course.created_at.strftime("%m/%d")
             ]
        )
    return render_template('auth/file_du/download.html',
                           courseManageLabels=courseManageLabels,
                           courseContent=courseContent
                           )


@auth.route('/acc', methods=['GET', 'POST'])
def acc(lend=None):
    account_form = AccountForm()
    if account_form.validate_on_submit():
        if account_form.modify_account_id.data.isdigit():
            modify_account_fuc(account_form)
        else:
            insert_account_fuc(account_form)
    else:
        account_form.account_date.data = datetime.now().strftime('%Y/%m/%d')
    del_accout_form = DelAccountForm()
    if del_accout_form.validate_on_submit():
        account_del = Account.query.filter(Account.id == del_accout_form.account_id.data).first()
        try:
            db.session.delete(account_del)
            flash(account_del.show_name + '已删除')
        except:
            flash(account_del.show_name + '删除失败')
    quartReportLabels = ['ID', '项目', '收入', '支出', '方式', '时间']
    dt = datetime.now()
    current_term = dt.strftime('%Y%m')
    query_term = request.args.get('query_term', current_term, type=str)
    query_year = request.args.get('query_year', None, type=bool)
    dt_cur_month = query_term_to_datetime(query_term)
    if query_year:
        dt_end = datetime(dt_cur_month.year + 1, 1, 1)
    else:
        dt_end = dt_cur_month + relativedelta(months=1)
    if lend:
        lend_query = Account.query.filter_by(show_name='姐姐还钱').order_by(Account.id.desc()).all()
        quartReportContent = quartReportPayShow(lend_query)
    else:
        if not current_user.is_authenticated:
            return current_app.login_manager.unauthorized()
        quartReportContent = quartReportPayShow(query_by_date(dt_cur_month, dt_end))
    quartReportContent.append(['', '共计', calSummary(quartReportContent, 2),
                               calSummary(quartReportContent, 3),
                               calSummary(quartReportContent, 2) - calSummary(quartReportContent, 3),
                               calAllSummary(quartReportContent, 2) - calAllSummary(quartReportContent, 3)])
    if lend is None:
        quartReportContent.append(
            ['', '结余', db.session.query(func.sum(Account.fee)).
                filter(Account.created_at < dt_end).first()[0]])
        item_list = ['招商卡', '支付宝', '微信']
        for i in item_list:
            quartReportContent.append(
                ['', f'{i}结余', db.session.query(func.sum(Account.fee)).
                    filter(Account.created_at < dt_end, Account.pay_type == i).first()[0]])
    return render_template('auth/quart_report.html',
                           lend=lend,
                           form=account_form,
                           query_terms=[current_term, (dt_cur_month - relativedelta(months=1)).strftime('%Y%m'),
                                        (dt_cur_month + relativedelta(months=1)).strftime('%Y%m'), f'{dt.year}01',
                                        query_term[0:4] + ' 全年' if query_year else query_term],
                           del_form=del_accout_form,
                           quartReportLabels=quartReportLabels,
                           last_record_time=get_last_record_time(),
                           quartReport=quartReportContent)


class Account(db.Model):
    __tablename__ = 'accounts'
    id = db.Column(db.Integer, primary_key=True)
    show_name = db.Column(db.Text())
    pay_type = db.Column(db.Text())
    fee = db.Column(db.Integer)
    refund_fee = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.now)
    last_updated_at = db.Column(db.DateTime(), default=datetime.now)

    def __init__(self, **kwargs):
        super(Account, self).__init__(**kwargs)


def query_by_date(dt_cur_month, dt_end):
    return Account.query.filter(Account.created_at >= dt_cur_month,
                                Account.created_at < dt_end). \
        order_by(Account.created_at.desc(), Account.id.desc()).all()


def quartReportPayShow(accounts_query):
    reportPay = []
    for account in accounts_query:
        if account.fee < 0:
            reportPay.append([
                account.id,
                account.show_name,
                '',
                account.refund_fee,
                account.pay_type,
                account.created_at.strftime("%m/%d")
            ])
        else:
            reportPay.append([
                account.id,
                account.show_name,
                account.fee,
                '',
                account.pay_type,
                account.created_at.strftime("%m/%d")
            ])
    return reportPay


def calSummary(content, i):
    summary = 0
    disable_set = ['转入', '转入微信', '微信转入', '招商卡结余', '微信结余', '余额宝-房租', '分成', '转出', '转入薪金宝']
    for item in content:
        if item[i] == '' or item[1] in disable_set:
            continue
        summary = summary + item[i]
    return summary


def calAllSummary(content, i):
    summary = 0
    for item in content:
        if item[i] == '':
            continue
        summary = summary + item[i]
    return summary


@auth.route('/institution', methods=['GET', 'POST'])
@login_required
def institution():
    current_term = current_app.config['CURRENT_TERM']
    if current_user.is_administrator() is False:
        return redirect(url_for('main.index'))
    query_term = request.args.get('query_term', current_term, type=str)
    add_institution_form = AddInstitutionForm()
    if add_institution_form.validate_on_submit():
        flash(add_institution_form.institution_name.data + '已添加')
        add_institution_infos_fuc(add_institution_form.institution_name.data,
                                  add_institution_form.institution_url.data,
                                  add_institution_form.job_category.data)
    forms = [add_institution_form]
    query_institution_infos = InstitutionInfo.query.filter_by(institution_period=query_term, disabled=False).all()
    statusLabels = ['编号', '招聘名称', '爬虫地址', '报考类别']
    return render_template('auth/institution.html',
                           statuslabels=statusLabels,
                           query_institution_infos=query_institution_infos,
                           query_term=query_term,
                           forms=forms)


@auth.route('/institutionManage', methods=['GET', 'POST'])
def institution_manage():
    query_institution_id = request.args.get('query_institution', 1, type=int)
    query_institution_info = InstitutionInfo.query.filter_by(institution_id=query_institution_id).first()
    statuslabels = ['序号', '部门名称', '职位名称', '招聘人数', '报名成功', '审核通过', '未审核', '总人数']
    form = UploadCoursesStus()
    form.form_body = '上传需为Excel表格，第一列为岗位代码，第二列为岗位类别，第三列为招聘人数，无表头'
    form.form_title = '导入岗位信息'
    if form.validate_on_submit():
        save_filename = os.path.join(current_app.config['BASE_DIR'], "Upload", form.file.data.filename)
        form.file.data.save(save_filename)
        wb = load_workbook(save_filename)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for row in ws.rows:
            institution_job_info = InstitutionJobInfo.query. \
                filter_by(institution_id=query_institution_id,
                          job_id=row[0].value).first()
            if institution_job_info is not None:
                institution_job_info.job_category = row[1].value
                institution_job_info.job_num = row[2].value
                institution_job_info.education = row[3].value
                institution_job_info.target = row[4].value
                db.session.add(institution_job_info)
                try:
                    db.session.commit()
                except:
                    db.session.rollback()
    forms = [form]
    query_institution_job_infos = InstitutionJobInfo.query.filter_by(institution_id=query_institution_id,
                                                                     job_category=query_institution_info.job_category). \
        order_by((InstitutionJobInfo.succeeded / InstitutionJobInfo.job_num).desc()).all()

    return render_template('auth/institution_manage.html',
                           statuslabels=statuslabels,
                           query_institution_info=query_institution_info,
                           job_infos=query_institution_job_infos,
                           forms=forms)


@auth.route('/spiderInstitution', methods=['GET', 'POST'])
def spider_institution():
    spider_institution_id = request.args.get('spider_institution_id', 1, type=int)
    query_institution_info = InstitutionInfo.query.filter_by(institution_id=spider_institution_id).first()
    spider_institution_jobs_fuc(query_institution_info)
    return redirect(url_for('auth.institution_manage', query_institution=spider_institution_id))


def query_term_to_datetime(query_term):
    try:
        dt_cur_month = datetime(int(query_term[0:4]), int(query_term[4:]), 1)
    except:
        dt = datetime.now()
        dt_cur_month = datetime(dt.year, dt.month, 1)
    return dt_cur_month


def insert_account_fuc(account_form):
    account_last = Account.query.filter().order_by(Account.id.desc()).first()
    account_fee = eval(account_form.account_fee.data)
    if not account_form.income.data:
        account_fee = -account_fee
    if account_last.show_name == account_form.account_name.data \
            and account_fee == account_last.fee \
            and account_last.last_updated_at > datetime.now() - timedelta(seconds=40):
        flash('重复插入')
    else:
        account = Account(
            show_name=account_form.account_name.data,
            pay_type=account_form.pay_type.data,
            fee=account_fee
        )
        if not account_form.income.data:
            account.refund_fee = -account_fee
        account.created_at = datetime.strptime(account_form.account_date.data, "%Y/%m/%d")
        db.session.add(account)
        try:
            if account.show_name == '转入微信':
                account_form.account_name.data = ''
                account_form.pay_type.data = '微信'
                account_new = Account(
                    show_name=account.show_name,
                    pay_type='微信',
                    fee=-account_fee,
                    refund_fee=account_fee,
                    created_at=account.created_at,
                )
                db.session.add(account_new)
            db.session.commit()
            flash('该账务已插入')
        except Exception as e:
            print(e)
            db.session.rollback()
            flash('该账务插入失败')


def modify_account_fuc(account_form):
    modify_account = Account.query.filter(Account.id == int(account_form.modify_account_id.data)).first()
    if modify_account is None:
        flash('该账务不存在')
        return
    if modify_account.show_name != account_form.account_name.data:
        flash(f'账务名从{modify_account.show_name}变为{account_form.account_name.data}')
        modify_account.show_name = account_form.account_name.data
    modify_account.pay_type = account_form.pay_type.data
    account_fee = eval(account_form.account_fee.data)
    if not account_form.income.data:
        account_fee = -account_fee
    if modify_account.fee != account_fee:
        flash(f'金额从{modify_account.fee}变为{account_fee}')
        modify_account.fee = account_fee
        if not account_form.income.data:
            modify_account.refund_fee = -account_fee
    db.session.add(modify_account)
    db.session.commit()
    flash('账务更新成功')


def get_last_record_time():
    last_account = Account.query.filter_by().order_by(Account.id.desc()).first()
    return last_account.last_updated_at.strftime('%Y/%m/%d %H:%M')


@auth.route('/buy', methods=['GET', 'POST'])
def buy():
    buy_form = BuyForm()
    del_buy_form = DelBuyForm()
    if buy_form.validate_on_submit():
        add_buy_info(buy_form)
    if del_buy_form.validate_on_submit():
        del_buy_info(del_buy_form.buy_info_id.data)
    return render_template('auth/buy.html',
                           buyInfoLabels=['编号', '商品名', '记录日期', '优先级', '备注'],
                           form=buy_form,
                           del_form=del_buy_form,
                           buy_infos=get_buy_infos())


def del_buy_info(del_buy_info_id):
    buy_info_del = BuyInfo.query.filter(BuyInfo.id == del_buy_info_id).first()
    try:
        db.session.delete(buy_info_del)
        flash(buy_info_del.buy_name + '已删除')
    except:
        flash(buy_info_del.buy_name + '删除失败')


def add_buy_info(buy_form):
    if buy_form.modify_account_id.data.isdigit():
        buy_info = BuyInfo.query.filter(BuyInfo.id == buy_form.modify_account_id.data).first()
    else:
        buy_info = BuyInfo()
    buy_info.buy_name = buy_form.good_name.data
    buy_info.priority_status = buy_form.good_priority.data
    buy_info.note = buy_form.note.data
    db.session.add(buy_info)


def get_buy_infos():
    return BuyInfo.query.filter().order_by(BuyInfo.status,
                                           BuyInfo.priority_status.desc()).all()


# @auth.route('/lend', methods=['GET', 'POST'])
# def lend():
#     return acc(lend=True)
